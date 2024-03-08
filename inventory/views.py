# inventory/views.py
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.db.models import F
from inventory.forms import UserRegistry, ProductForm, OrderForm
from inventory.models import Product, Order
from django.db.models import Sum
from django.http import HttpResponse
import openpyxl, datetime


@login_required
def index(request):
    orders_user = Order.objects.all()
    users = User.objects.all()[:2]
    orders_adm = Order.objects.all()[:2]
    products = Product.objects.all()[:2]
    reg_users = len(User.objects.all())
    all_prods = len(Product.objects.all())
    all_orders = len(Order.objects.all())
    context = {
        "title": "Home",
        "orders": orders_user,
        "orders_adm": orders_adm,
        "users": users,
        "products": products,
        "count_users": reg_users,
        "count_products": all_prods,
        "count_orders": all_orders,
    }
    return render(request, "inventory/index.html", context)

def index2(request):
    context = {}
    return render(request, "inventory/index2.html", context)


def about(request):
    context = {}
    return render(request, "inventory/about.html", context)

def contact(request):
    context = {}
    return render(request, "inventory/contact.html", context)

@login_required
def products(request):
    products = Product.objects.all()
    search = request.GET.get("search", "")
    if search:
        products = products.filter(name__icontains=search).order_by('-id')
    if request.method == "POST":
        form = ProductForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect("products")
    else:
        form = ProductForm()

        prices = {product.name: product.price for product in products}

        context = {
            "title": "Products", "products": products, "form": form, "prices": prices,
            "search": search,
        }
        return render(request, "inventory/products.html", context)

@login_required
def orders(request):
    orders = Order.objects.all()
    print([i for i in request])
    if request.method == "POST":
        form = OrderForm(request.POST)
        if form.is_valid():
            instance = form.save(commit=False)
            instance.created_by = request.user
            instance.save()
            return redirect("orders")
    else:
        form = OrderForm()
    context = {"title": "Orders", "orders": orders, "form": form}
    return render(request, "inventory/orders.html", context)

@login_required
def users(request):
    users = User.objects.all()
    context = {"title": "Users", "users": users}
    return render(request, "inventory/users.html", context)

@login_required
def user(request):
    context = {"profile": "User Profile"}
    return render(request, "inventory/user.html", context)

def register(request):
    if request.method == "POST":
        form = UserRegistry(request.POST)
        if form.is_valid():
            form.save()
            return redirect("login")
    else:
        form = UserRegistry()
    context = {"register": "Register", "form": form}
    return render(request, "inventory/register.html", context)

def completed_orders():
    return Order.objects.filter(status='completed')

def cancelled_orders():
    return Order.objects.filter(status='cancelled')

@login_required
def sales_report(request):
    orders_by_date = Order.objects.order_by('date')
    total_sales = sum(order.get_total() for order in orders_by_date)  # Calculate total sales
    
    context = {
        "title": "Sales Report",
        "orders_by_date": orders_by_date,
        "total_sales": total_sales

    }
    if request.method == "POST" and request.POST.get("download_excel"):
        return generate_sales_report(request, total_sales)
    else:
        # Return an HttpResponse if the request method is not POST
        return render(request, "inventory/sales_report.html", context)

def generate_sales_report(request, total_sales):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="salesreport.xlsx"'

    workbook = openpyxl.Workbook()
    workbook.iso_dates = True
    worksheet = workbook.active
    worksheet.title = 'Dawa Pharmacy Sales Report'

    # Write header row
    header = ['Order ID', 'Product', 'Quantity', 'Price (Ksh)', 'Total (Ksh)', 'Date', 'Status']
    for col_num, column_title in enumerate(header, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = column_title

    # Write data rows
    queryset = Order.objects.all().annotate(
        total=F('order_quantity')*F('product__price')
        ).values_list("id", "product__name", "order_quantity", "product__price", "total", "date" ,"status")
    for row_num, row in enumerate(queryset, 1):
        for col_num, cell_value in enumerate(row, 1):
            if col_num == 6:
                cell_value = cell_value.replace(tzinfo=None)
                cell_value = cell_value.strftime("%b %d, %Y, %I:%M:%p")
            cell = worksheet.cell(row=row_num+1, column=col_num)
            cell.value = cell_value
            
    workbook.save(response)

    return response 

