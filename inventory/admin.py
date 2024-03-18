import io
from django.contrib import admin
from django.conf import settings
from django.core.mail import send_mail
from django.db.models.query import QuerySet
from django.http import HttpResponse
import openpyxl
from django.template.loader import render_to_string
from weasyprint import HTML
from datetime import datetime
from django.utils import timezone
from django.db.models import Q

# Register your models here.
# inventory/admin.py
from inventory.models import Product, Order, UserProfile, SalesInvoice

admin.site.site_header = "Inventory Admin"

class ProductAdmin(admin.ModelAdmin):
    model = Product
    list_display = ("name", "category", "quantity", "price", "reorder_date")
    list_filter = ["category"]
    search_fields = ["name"]
    list_editable = ["quantity", "price"]
    actions = ["reorder_product_from_supplier"]

    @admin.action(description="Reorder products that are below minimum stock level")
    def reorder_product_from_supplier(self, request, queryset):
        message = ''
        current_datetime = timezone.localtime(timezone.now())

        for product in queryset:
            if product.quantity <= settings.MIN_STOCK_QUANTITY:
                reorder_quantity = 10
                print(f"Reordering {reorder_quantity} units of {product.name}")
                product.quantity += reorder_quantity
                product.reorder_date = current_datetime
                product.save()
                print(f"Reordered {reorder_quantity} units of {product.name}")
                message += f"{product.name}: {reorder_quantity}.\n"

        if message:
            message = "Please supply me with the requested quantities for the following products by 8:00 PM today:\n" + message
            send_mail(
                subject="REQUEST FOR PRODUCTS SUPPLY",
                message=message,
                from_email=settings.EMAIL_HOST_USER,
                recipient_list=[settings.EMAIL_SUPPLIER],
                fail_silently=False,
            )
            self.message_user(request, "Reorder notifications sent to the supplier.")

        else:
            self.message_user(request, "No products require reordering.")

class InputFilter(admin.SimpleListFilter):
    template = 'admin/input_filter.html'

    def queryset(self, request, queryset):
        pass

    def lookups(self, request, model_admin):
        # Dummy, required to show the filter.
        return ((),)

    def choices(self, changelist):
        # Grab only the "all" option.
        all_choice = next(super().choices(changelist))
        all_choice['query_parts'] = (
            (k, v)
            for k, v in changelist.get_filters_params().items()
            if k != self.parameter_name
        )
        yield all_choice


class ProductFilter(InputFilter):
    parameter_name = 'product'
    title = 'Product'

    def queryset(self, request, queryset):
        term = self.value()

        if term is None:
            return

        product = Q()
        for bit in term:
            product &= (
                Q(product__name__icontains=bit)
            )

        return queryset.filter(product)

class ClientFilter(InputFilter):
    parameter_name = 'client'
    title = 'Client'

    def queryset(self, request, queryset):
        term = self.value()

        if term is None:
            return

        product = Q()
        for bit in term:
            client &= (
                Q(client__name__icontains=bit)
            )

        return queryset.filter(client)

class OrderAdmin(admin.ModelAdmin):
    model = Order
    list_display = ("product", "created_by", "order_quantity", "date", "client")
    list_filter = ["date", "product", "client", ClientFilter, ProductFilter]
    search_fields = ["product"]
    actions = ["download_excel_report", "download_pdf_report"]

    @admin.action(description="Download PDF Report")
    def download_pdf_report(self, request, queryset):
        total_sales = sum(order.get_total() for order in queryset)
        context = {
            "title": "Sales Report",
            "orders": queryset,
            "total_sales": total_sales
        }
        html = render_to_string('inventory/report.html', context)
        with io.BytesIO() as buf:
            buf.name = "salesreport.pdf"
            HTML(string=html).write_pdf(buf)
            response = HttpResponse(buf.getvalue(), content_type='application/pdf')
            response['Content-Disposition'] = 'attachment; filename="salesreport.pdf"'

        return response

    @admin.action(description="Download Excel Report")
    def download_excel_report(self, request, queryset):
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="salesreport.xlsx"'

        workbook = openpyxl.Workbook()
        workbook.iso_dates = True
        worksheet = workbook.active
        worksheet.title = 'Dawa Pharmacy Sales Report'

        # Write header row
        header = ['Product', 'Created_by', 'Order quantity', 'Date', "Client"]
        for col_num, column_title in enumerate(header, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = column_title

        queryset = queryset.values_list("product__name", "created_by__username", "order_quantity", "date", "client")
        row_num = 2
        
        for row in queryset:
            for col_num, cell_value in enumerate(row, 1):
                if col_num == 4:
                    cell_value = cell_value.replace(tzinfo=None)
                    cell_value = cell_value.strftime("%b %d, %Y, %I:%M:%p")
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value

            row_num += 1
                
        workbook.save(response)
        return response
    
class UserProfileAdmin(admin.ModelAdmin):
    model = UserProfile
    list_display = ("user", "physical_address", "mobile", "picture")
    # list_filter = ["user__username"]
    search_fields = ["user__username"]

admin.site.register(Product, ProductAdmin)
admin.site.register(Order, OrderAdmin)
admin.site.register(UserProfile, UserProfileAdmin) 
admin.site.register(SalesInvoice)
